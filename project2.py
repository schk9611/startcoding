import requests
from bs4 import BeautifulSoup
import openpyxl
import os

# 엑셀 파일을 만들고자 하는 위치
fpath = '/Users/songyu/Documents/projects/python-projects/startcoding/data.xlsx'

# 엑셀 만들기(만약 동일한 위치에 동일한 이름의 파일이 존재하면 실행하지 않음)
if os.path.isdir(fpath) == True:
    pass
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '종목'
    ws['B1'] = '현재가'
    ws['A2'] = '삼성전자'
    ws['A3'] = 'SK하이닉스'
    ws['A4'] = '카카오'

    wb.save('data.xlsx')

wb = openpyxl.load_workbook(fpath)
ws = wb.active  # 현재 활성화된 시트 선택

# 종목 코드 리스트
codes = [
    '005930',
    '000660',
    '035720'
]

row = 2
for code in codes:
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    price = soup.select_one("#_nowVal").text
    price = price.replace(',', '')
    print(price)
    ws[f'B{row}'] = int(price)
    row = row + 1

wb.save(fpath)
